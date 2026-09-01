import io
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill

from .. import models
from ..database import get_db

router = APIRouter(prefix="/dashboard", tags=["Dashboard"])


def _achievement_counts(db: Session):
    return dict(
        db.query(models.Achievement.student_id, func.count(models.Achievement.id))
        .group_by(models.Achievement.student_id).all()
    )


@router.get("/stats")
def dashboard_stats(year: Optional[str] = None, section: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(models.Student)
    if year:
        query = query.filter(models.Student.year == year)
    if section:
        query = query.filter(models.Student.section == section.upper())
    students = query.all()
    counts = _achievement_counts(db)

    latest_event_name_by_student = {}
    for achievement in db.query(models.Achievement).order_by(models.Achievement.student_id, models.Achievement.created_at.desc(), models.Achievement.id.desc()):
        student_id = achievement.student_id
        if student_id not in latest_event_name_by_student:
            latest_event_name_by_student[student_id] = achievement.event_name

    top_performers = sorted(
        [{
            "id": s.id,
            "name": f"{s.first_name} {s.last_name or ''}".strip(),
            "roll_no": s.roll_no,
            "section": s.section,
            "year": s.year,
            "achievement_count": counts.get(s.id, 0),
            "event_name": latest_event_name_by_student.get(s.id) or "—",
        } for s in students],
        key=lambda x: x["achievement_count"], reverse=True
    )
    participants = [p for p in top_performers if p["achievement_count"] > 0]
    non_participants = [p for p in top_performers if p["achievement_count"] == 0]

    return {
        "total_students": len(students),
        "total_participants": len(participants),
        "total_non_participants": len(non_participants),
        "top_performers": top_performers[:20],
        "participants": participants,
        "non_participants": non_participants,
    }


def _export_list(items, filename, cols, extractor):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="4B3F72", end_color="4B3F72", fill_type="solid")
    for i, item in enumerate(items, start=1):
        ws.append(extractor(i, item))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/top-performers")
def export_top_performers(year: Optional[str] = None, section: Optional[str] = None, db: Session = Depends(get_db)):
    data = dashboard_stats(year, section, db)
    return _export_list(
        data["top_performers"], "top_performers.xlsx",
        ["S No", "Name", "Roll No", "Section", "Year", "Event Name", "Achievements"],
        lambda i, p: [i, p["name"], p["roll_no"], p["section"], p["year"], p["event_name"], p["achievement_count"]],
    )


@router.get("/export/participants")
def export_participants(year: Optional[str] = None, section: Optional[str] = None, db: Session = Depends(get_db)):
    data = dashboard_stats(year, section, db)
    return _export_list(
        data["participants"], "participants.xlsx",
        ["S No", "Name", "Roll No", "Section", "Year", "Event Name", "Achievements"],
        lambda i, p: [i, p["name"], p["roll_no"], p["section"], p["year"], p["event_name"], p["achievement_count"]],
    )


@router.get("/export/non-participants")
def export_non_participants(year: Optional[str] = None, section: Optional[str] = None, db: Session = Depends(get_db)):
    data = dashboard_stats(year, section, db)
    return _export_list(
        data["non_participants"], "non_participants.xlsx",
        ["S No", "Name", "Roll No", "Section", "Year", "Event Name"],
        lambda i, p: [i, p["name"], p["roll_no"], p["section"], p["year"], p["event_name"]],
    )
