import io
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill

from .. import models
from ..database import get_db

router = APIRouter(prefix="/events", tags=["Event Search"])


def _search(db: Session, event_name: Optional[str], event_type: Optional[str]):
    query = db.query(models.Achievement).options(joinedload(models.Achievement.student))
    if event_name:
        query = query.filter(models.Achievement.event_name.ilike(f"%{event_name}%"))
    if event_type:
        query = query.filter(models.Achievement.event_type == event_type)
    return query.order_by(models.Achievement.event_date.desc()).all()


@router.get("/search")
def search_events(event_name: Optional[str] = None, event_type: Optional[str] = None, db: Session = Depends(get_db)):
    results = _search(db, event_name, event_type)
    return [
        {
            "achievement_id": a.id,
            "student_id": a.student.id,
            "student_name": f"{a.student.first_name} {a.student.last_name or ''}".strip(),
            "roll_no": a.student.roll_no,
            "event_name": a.event_name,
            "event_type": a.event_type,
            "prize_type": a.prize_type,
            "event_date": a.event_date,
        }
        for a in results
    ]


@router.get("/search/export")
def export_search(event_name: Optional[str] = None, event_type: Optional[str] = None, db: Session = Depends(get_db)):
    results = _search(db, event_name, event_type)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Event Results"
    headers = ["S No", "Name", "Roll No", "Event Name", "Event Type", "Prize", "Participation Date"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="4B3F72", end_color="4B3F72", fill_type="solid")

    for i, a in enumerate(results, start=1):
        ws.append([
            i, f"{a.student.first_name} {a.student.last_name or ''}".strip(), a.student.roll_no,
            a.event_name, a.event_type or "", a.prize_type or "", a.event_date or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=event_search_results.xlsx"},
    )
