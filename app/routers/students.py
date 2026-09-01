import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, func
from typing import Optional, List
import openpyxl
from openpyxl.styles import Font, PatternFill

from .. import models, schemas
from ..database import get_db
from ..auth import get_current_admin, get_current_student_or_admin

router = APIRouter(prefix="/students", tags=["Students"])


@router.get("/", response_model=List[schemas.StudentOut])
def list_students(
    year: Optional[str] = None,
    section: Optional[str] = None,
    department: Optional[str] = None,
    name: Optional[str] = None,
    db: Session = Depends(get_db),
):
    query = db.query(models.Student)
    if year:
        query = query.filter(models.Student.year == year)
    if section:
        query = query.filter(models.Student.section == section.upper())
    if department:
        query = query.filter(models.Student.department == department)
    if name:
        like = f"%{name}%"
        query = query.filter(
            or_(
                models.Student.roll_no.ilike(like),
                models.Student.first_name.ilike(like),
                models.Student.last_name.ilike(like),
                models.Student.email.ilike(like),
            )
        )
    students = query.order_by(models.Student.roll_no).all()

    counts = dict(
        db.query(models.Achievement.student_id, func.count(models.Achievement.id))
        .group_by(models.Achievement.student_id).all()
    )
    out = []
    for s in students:
        d = schemas.StudentOut.model_validate(s).model_dump()
        d["achievement_count"] = counts.get(s.id, 0)
        out.append(d)
    return out


@router.get("/export")
def export_students(
    year: Optional[str] = None,
    section: Optional[str] = None,
    db: Session = Depends(get_db),
):
    query = db.query(models.Student)
    if year:
        query = query.filter(models.Student.year == year)
    if section:
        query = query.filter(models.Student.section == section.upper())
    students = query.order_by(models.Student.roll_no).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    headers = ["S No", "Roll No", "Reg No", "Name", "Year", "Section", "Dept", "Email", "Mobile"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="4B3F72", end_color="4B3F72", fill_type="solid")

    for i, s in enumerate(students, start=1):
        ws.append([
            i, s.roll_no, s.reg_no or "", f"{s.first_name} {s.last_name or ''}".strip(),
            s.year, s.section, s.department, s.email or "", s.mobile_number or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=students_export.xlsx"},
    )


@router.get("/{student_id}", response_model=schemas.StudentDetail)
def get_student(student_id: int, db: Session = Depends(get_db)):
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    return student


@router.get("/{student_id}/export")
def export_single_student(student_id: int, db: Session = Depends(get_db)):
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profile"
    ws.append(["Field", "Value"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="4B3F72", end_color="4B3F72", fill_type="solid")

    ws.append(["Name", f"{student.first_name} {student.last_name or ''}".strip()])
    ws.append(["Roll No", student.roll_no])
    ws.append(["Reg No", student.reg_no or ""])
    ws.append(["Year", student.year])
    ws.append(["Section", student.section])
    ws.append(["Department", student.department])
    ws.append(["Email", student.email or ""])
    ws.append(["Mobile", student.mobile_number or ""])

    ws2 = wb.create_sheet("Achievements")
    ws2.append(["S No", "Event Name", "Event Type", "Prize", "Date", "Organizer"])
    for c in ws2[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="2E7D5B", end_color="2E7D5B", fill_type="solid")
    for i, a in enumerate(student.achievements, start=1):
        ws2.append([i, a.event_name, a.event_type or "", a.prize_type or "", a.event_date or "", a.organizer or ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{student.roll_no}_profile.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.post("/", response_model=schemas.StudentOut)
def create_student(
    student: schemas.StudentCreate,
    db: Session = Depends(get_db),
    admin: str = Depends(get_current_admin),
):
    db_student = models.Student(**student.model_dump())
    db.add(db_student)
    db.commit()
    db.refresh(db_student)
    return db_student


@router.put("/{student_id}", response_model=schemas.StudentOut)
def update_student(
    student_id: int,
    student: schemas.StudentCreate,
    db: Session = Depends(get_db),
    admin: str = Depends(get_current_admin),
):
    db_student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not db_student:
        raise HTTPException(status_code=404, detail="Student not found")
    for key, value in student.model_dump().items():
        setattr(db_student, key, value)
    db.commit()
    db.refresh(db_student)
    return db_student


@router.delete("/bulk/by-class")
def delete_students_by_class(
    year: str,
    section: str,
    db: Session = Depends(get_db),
    admin: str = Depends(get_current_admin),
):
    """Deletes ALL students (and their achievements) matching year + section."""
    query = db.query(models.Student).filter(
        models.Student.year == year,
        models.Student.section == section.upper(),
    )
    count = query.count()
    if count == 0:
        raise HTTPException(status_code=404, detail="No students match this year/section")

    for student in query.all():
        db.delete(student)
    db.commit()

    return {"message": f"Deleted {count} students", "deleted_count": count}


@router.delete("/{student_id}")
def delete_student(
    student_id: int,
    db: Session = Depends(get_db),
    admin: str = Depends(get_current_admin),
):
    db_student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not db_student:
        raise HTTPException(status_code=404, detail="Student not found")
    db.delete(db_student)
    db.commit()
    return {"message": "Student deleted"}

@router.post("/{student_id}/upload-photo")
def upload_student_photo(
    student_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_student_or_admin),
):
    from ..storage import upload_photo_bytes
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    if current_user.get("role") == "student" and current_user.get("student_id") != student_id:
        raise HTTPException(status_code=403, detail="You can upload only your own photo")

    ext = os.path.splitext(file.filename)[1] or ".jpg"
    file_bytes = file.file.read()
    public_url = upload_photo_bytes(file_bytes, f"{student.roll_no}{ext}")

    student.photo_path = public_url
    db.commit()
    return {"message": "Photo uploaded", "photo_path": public_url}
